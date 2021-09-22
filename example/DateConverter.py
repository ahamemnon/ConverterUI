import openpyxl
import os
from datetime import date

class DateConverter:
    def __init__(self, fileName):
        self.fileName = fileName
    
    def convertDate(self, val):
        """
        конвертирует текстовую строку формата yyyymmdd в дату
        """
        val = str(val)
        val = date(int(val[:4]), int(val[4:6]), int(val[6:]))
        return val
    
    def processFile(self):
        self.wb = openpyxl.load_workbook(self.fileName)
        self.ws = self.wb.worksheets[0]
        
        for i in range(2, self.ws.max_row + 1):
            for col in [9, 10]:
                self.ws.cell(row=i, column=col).number_format = 'dd-mm-yyyy'
                self.ws.cell(row=i, column=col).value = self.convertDate(self.ws.cell(row=i, column=col).value)
        
    def writeFile(self):
        fileName, fileExtension = os.path.splitext(self.fileName)
        newFileName = fileName + '_conv' + fileExtension
        self.wb.save(newFileName)

if __name__ == '__main__':
    files = ['1.xlsx',
             '2.xlsx']
    
    for file in files:
        dc = DateConverter(file)
        try:
            dc.processFile()
            dc.writeFile()
        except FileNotFoundError:
            print('File Not Found')
        except PermissionError:
            print('Permission Error')
        else:
            print(file)